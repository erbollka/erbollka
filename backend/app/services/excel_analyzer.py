import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from statistics import median, pstdev
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter


TOTAL_PATTERNS = ("total", "итого", "сумма", "всего")
NEGATIVE_FORBIDDEN_PATTERNS = ("sale", "продаж", "salary", "зарп", "bonus", "прем", "остат", "expense", "расход")
REQUIRED_PATTERNS = ("date", "дата", "employee", "сотруд", "sales", "продаж", "salary", "зарп", "total", "итого")
METRIC_PATTERNS = {
    "sales": ("sale", "revenue", "выруч", "продаж"),
    "salary": ("salary", "payroll", "зарп", "оклад", "фот"),
    "bonus": ("bonus", "прем"),
    "expense": ("expense", "расход", "затрат"),
    "profit": ("profit", "приб", "марж"),
    "discount": ("discount", "скид"),
    "return": ("return", "refund", "возврат"),
}


@dataclass
class Finding:
    code: str
    severity: str
    sheet_name: str
    title: str
    description: str
    suggested_fix: str
    cell: str | None = None
    row_number: int | None = None
    column_name: str | None = None
    evidence: dict[str, Any] = field(default_factory=dict)
    confidence: float = 0.85


@dataclass
class WorkbookAudit:
    file_sha256: str
    workbook_profile: dict[str, Any]
    findings: list[Finding]
    risk_score: int
    risk_level: str
    summary: str
    recommendations: list[str]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def analyze_workbook(path: Path) -> WorkbookAudit:
    findings: list[Finding] = []
    workbook_profile = _profile_workbook(path)
    workbook_formula = openpyxl.load_workbook(path, data_only=False, read_only=False)
    workbook_values = openpyxl.load_workbook(path, data_only=True, read_only=False)

    for sheet_name in workbook_formula.sheetnames:
        ws_formula = workbook_formula[sheet_name]
        ws_values = workbook_values[sheet_name]
        findings.extend(_find_formula_issues(sheet_name, ws_formula, ws_values))
        findings.extend(_find_required_blanks(sheet_name, ws_values))
        findings.extend(_find_duplicate_rows(path, sheet_name))
        findings.extend(_find_negative_values(sheet_name, ws_values))
        findings.extend(_find_outliers(path, sheet_name))
        findings.extend(_find_total_mismatches(sheet_name, ws_values))
        findings.extend(_find_business_logic_issues(path, sheet_name))

    findings.extend(_find_workbook_business_issues(workbook_profile))
    risk_score = _risk_score(findings)
    risk_level = _risk_level(risk_score)
    quality_score = max(0, 100 - risk_score)
    workbook_profile["quality_score"] = quality_score
    return WorkbookAudit(
        file_sha256=sha256_file(path),
        workbook_profile=workbook_profile,
        findings=findings,
        risk_score=risk_score,
        risk_level=risk_level,
        summary=_summary(findings, risk_level, quality_score),
        recommendations=_recommendations(findings),
    )


def _profile_workbook(path: Path) -> dict[str, Any]:
    excel = pd.ExcelFile(path)
    sheets = []
    totals = {key: 0.0 for key in METRIC_PATTERNS}
    for sheet_name in excel.sheet_names:
        frame = excel.parse(sheet_name=sheet_name, header=None, nrows=80)
        parsed = excel.parse(sheet_name=sheet_name)
        numeric_frame = frame.select_dtypes(include=[np.number])
        sheet_totals = _metric_totals(parsed)
        for key, value in sheet_totals.items():
            totals[key] += value
        sheets.append(
            {
                "name": sheet_name,
                "rows_sampled": int(frame.shape[0]),
                "columns_sampled": int(frame.shape[1]),
                "non_empty_cells_sampled": int(frame.notna().sum().sum()),
                "numeric_columns_sampled": int(numeric_frame.shape[1]),
                "metric_totals": sheet_totals,
            }
        )
    return {"sheet_count": len(sheets), "sheets": sheets, "metric_totals": {k: round(v, 2) for k, v in totals.items()}}


def _metric_totals(frame: pd.DataFrame) -> dict[str, float]:
    totals = {key: 0.0 for key in METRIC_PATTERNS}
    for column in frame.columns:
        normalized = str(column).lower()
        series = pd.to_numeric(frame[column], errors="coerce").dropna()
        if series.empty:
            continue
        for metric, patterns in METRIC_PATTERNS.items():
            if any(pattern in normalized for pattern in patterns):
                totals[metric] += float(series.sum())
    return {key: round(value, 2) for key, value in totals.items()}


def _find_formula_issues(sheet_name: str, ws_formula: Any, ws_values: Any) -> list[Finding]:
    findings: list[Finding] = []
    for row in ws_formula.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue
            cached = ws_values[cell.coordinate].value
            formula = cell.value.upper()
            if cached is None:
                findings.append(
                    Finding(
                        code="FORMULA_NO_CACHED_VALUE",
                        severity="medium",
                        sheet_name=sheet_name,
                        cell=cell.coordinate,
                        title="Формула без сохраненного результата",
                        description=f"В ячейке {cell.coordinate} есть формула, но в файле нет сохраненного вычисленного значения.",
                        suggested_fix="Откройте файл в Excel, пересчитайте книгу и сохраните ее перед повторной загрузкой.",
                        evidence={"formula": cell.value},
                    )
                )
            if "SUM(" in formula and _has_suspicious_sum_range(cell.value):
                findings.append(
                    Finding(
                        code="FORMULA_SUSPICIOUS_SUM_RANGE",
                        severity="high",
                        sheet_name=sheet_name,
                        cell=cell.coordinate,
                        title="Подозрительный диапазон суммы",
                        description=f"Формула в {cell.coordinate} может суммировать неполный или слишком широкий диапазон.",
                        suggested_fix="Проверьте, что в SUM попали все строки месяца и не попали итоговые строки.",
                        evidence={"formula": cell.value},
                    )
                )
    return findings[:80]


def _has_suspicious_sum_range(formula: str) -> bool:
    ranges = re.findall(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", formula.upper())
    for start_col, start_row, end_col, end_row in ranges:
        if start_col != end_col:
            return True
        if int(end_row) - int(start_row) > 400:
            return True
    return False


def _find_required_blanks(sheet_name: str, ws_values: Any) -> list[Finding]:
    findings: list[Finding] = []
    header_row, headers = _detect_headers(ws_values)
    for column_index, header in headers.items():
        normalized = str(header).lower()
        if not any(pattern in normalized for pattern in REQUIRED_PATTERNS):
            continue
        for row_index in range(header_row + 1, min(ws_values.max_row, 500) + 1):
            value = ws_values.cell(row=row_index, column=column_index).value
            if value in (None, ""):
                findings.append(
                    Finding(
                        code="REQUIRED_CELL_EMPTY",
                        severity="medium",
                        sheet_name=sheet_name,
                        cell=f"{get_column_letter(column_index)}{row_index}",
                        row_number=row_index,
                        column_name=str(header),
                        title="Пустая обязательная ячейка",
                        description=f"В колонке «{header}» пустая ячейка в строке {row_index}.",
                        suggested_fix="Заполните значение или удалите строку, если она не относится к отчету.",
                    )
                )
                if len(findings) >= 30:
                    return findings
    return findings


def _detect_headers(ws_values: Any) -> tuple[int, dict[int, str]]:
    best_row = 1
    best_score = -1
    for row_index in range(1, min(ws_values.max_row, 12) + 1):
        values = [ws_values.cell(row=row_index, column=col).value for col in range(1, ws_values.max_column + 1)]
        score = sum(1 for value in values if isinstance(value, str) and len(value.strip()) > 1)
        if score > best_score:
            best_row = row_index
            best_score = score
    headers: dict[int, str] = {}
    for col in range(1, ws_values.max_column + 1):
        value = ws_values.cell(row=best_row, column=col).value
        if value not in (None, ""):
            headers[col] = str(value)
    return best_row, headers


def _find_duplicate_rows(path: Path, sheet_name: str) -> list[Finding]:
    frame = pd.read_excel(path, sheet_name=sheet_name, dtype=str).dropna(how="all")
    if frame.empty:
        return []
    duplicated = frame.duplicated(keep=False)
    return [
        Finding(
            code="DUPLICATE_ROW",
            severity="low",
            sheet_name=sheet_name,
            row_number=int(index) + 2,
            title="Дубликат строки",
            description=f"Строка {int(index) + 2} полностью совпадает с другой строкой на листе.",
            suggested_fix="Проверьте, не была ли строка случайно скопирована дважды.",
        )
        for index in frame[duplicated].index[:20]
    ]


def _find_negative_values(sheet_name: str, ws_values: Any) -> list[Finding]:
    findings: list[Finding] = []
    _, headers = _detect_headers(ws_values)
    for column_index, header in headers.items():
        normalized = str(header).lower()
        if not any(pattern in normalized for pattern in NEGATIVE_FORBIDDEN_PATTERNS):
            continue
        for row_index in range(2, min(ws_values.max_row, 1000) + 1):
            value = ws_values.cell(row=row_index, column=column_index).value
            if isinstance(value, (int, float)) and value < 0:
                findings.append(
                    Finding(
                        code="NEGATIVE_VALUE",
                        severity="high",
                        sheet_name=sheet_name,
                        cell=f"{get_column_letter(column_index)}{row_index}",
                        row_number=row_index,
                        column_name=str(header),
                        title="Отрицательное значение",
                        description=f"В колонке «{header}» найдено отрицательное значение {value}.",
                        suggested_fix="Проверьте знак числа и корректность возвратов или сторно.",
                        evidence={"value": value},
                    )
                )
    return findings[:40]


def _find_outliers(path: Path, sheet_name: str) -> list[Finding]:
    frame = pd.read_excel(path, sheet_name=sheet_name)
    findings: list[Finding] = []
    numeric_columns = frame.select_dtypes(include=[np.number]).columns
    for column in numeric_columns:
        series = frame[column].dropna()
        if len(series) < 8:
            continue
        sigma = float(pstdev(series))
        if sigma == 0:
            continue
        center = float(median(series))
        for index, value in series.items():
            z_score = abs((float(value) - center) / sigma)
            if z_score >= 3:
                findings.append(
                    Finding(
                        code="NUMERIC_OUTLIER",
                        severity="medium",
                        sheet_name=sheet_name,
                        row_number=int(index) + 2,
                        column_name=str(column),
                        title="Аномальное значение",
                        description=f"Значение {value} в колонке «{column}» сильно отличается от остальных.",
                        suggested_fix="Сверьте сумму с первичным документом, POS-системой или ведомостью.",
                        evidence={"value": float(value), "median": center, "z_score": round(z_score, 2)},
                    )
                )
    return findings[:40]


def _find_total_mismatches(sheet_name: str, ws_values: Any) -> list[Finding]:
    findings: list[Finding] = []
    for row_index in range(1, ws_values.max_row + 1):
        row_values = [ws_values.cell(row=row_index, column=col).value for col in range(1, ws_values.max_column + 1)]
        row_text = " ".join(str(value).lower() for value in row_values if isinstance(value, str))
        if not any(pattern in row_text for pattern in TOTAL_PATTERNS):
            continue
        numeric_cells = [
            (col, value)
            for col, value in enumerate(row_values, start=1)
            if isinstance(value, (int, float)) and not isinstance(value, bool)
        ]
        for col, total_value in numeric_cells:
            values_above = [ws_values.cell(row=row, column=col).value for row in range(max(1, row_index - 40), row_index)]
            numeric_above = [float(value) for value in values_above if isinstance(value, (int, float))]
            if len(numeric_above) >= 2:
                calculated = round(sum(numeric_above), 2)
                expected = round(float(total_value), 2)
                if abs(calculated - expected) > max(1.0, abs(expected) * 0.01):
                    findings.append(
                        Finding(
                            code="TOTAL_MISMATCH",
                            severity="critical",
                            sheet_name=sheet_name,
                            cell=f"{get_column_letter(col)}{row_index}",
                            row_number=row_index,
                            title="Итог не сходится",
                            description=f"Итог {expected} не совпадает с суммой строк выше {calculated}.",
                            suggested_fix="Проверьте диапазон формулы и строки, которые входят в итог.",
                            evidence={"expected": expected, "calculated": calculated, "difference": round(expected - calculated, 2)},
                        )
                    )
    return findings[:30]


def _find_business_logic_issues(path: Path, sheet_name: str) -> list[Finding]:
    frame = pd.read_excel(path, sheet_name=sheet_name)
    if frame.empty:
        return []
    findings: list[Finding] = []
    columns = {str(column).lower(): column for column in frame.columns}
    salary_col = _find_column(columns, METRIC_PATTERNS["salary"])
    bonus_col = _find_column(columns, METRIC_PATTERNS["bonus"])
    discount_col = _find_column(columns, METRIC_PATTERNS["discount"])
    return_col = _find_column(columns, METRIC_PATTERNS["return"])
    date_col = _find_column(columns, ("date", "дата"))

    if salary_col is not None and bonus_col is not None:
        salary = pd.to_numeric(frame[salary_col], errors="coerce")
        bonus = pd.to_numeric(frame[bonus_col], errors="coerce")
        for index, (salary_value, bonus_value) in enumerate(zip(salary, bonus), start=2):
            if pd.notna(salary_value) and pd.notna(bonus_value) and salary_value > 0 and bonus_value > salary_value * 0.5:
                findings.append(
                    Finding(
                        code="BONUS_LIMIT_EXCEEDED",
                        severity="high",
                        sheet_name=sheet_name,
                        row_number=index,
                        column_name=str(bonus_col),
                        title="Премия выше лимита",
                        description=f"Премия {bonus_value:.2f} превышает 50% зарплаты {salary_value:.2f}.",
                        suggested_fix="Проверьте основание премии и согласование с бухгалтерией.",
                        evidence={"salary": float(salary_value), "bonus": float(bonus_value), "limit_ratio": 0.5},
                    )
                )

    if discount_col is not None:
        discounts = pd.to_numeric(frame[discount_col], errors="coerce")
        for index, value in discounts.dropna().items():
            if value > 50:
                findings.append(
                    Finding(
                        code="SUSPICIOUS_DISCOUNT",
                        severity="medium",
                        sheet_name=sheet_name,
                        row_number=int(index) + 2,
                        column_name=str(discount_col),
                        title="Подозрительно большая скидка",
                        description=f"Скидка {value:.2f} выглядит выше обычного лимита.",
                        suggested_fix="Проверьте разрешение на скидку и соответствие акции.",
                        evidence={"discount": float(value)},
                    )
                )

    if return_col is not None and date_col is not None:
        dates = pd.to_datetime(frame[date_col], errors="coerce")
        returns = pd.to_numeric(frame[return_col], errors="coerce").abs()
        return_median = float(returns.dropna().median()) if returns.notna().any() else 0.0
        for index, (date_value, return_value) in enumerate(zip(dates, returns), start=2):
            if pd.notna(date_value) and pd.notna(return_value) and date_value.day >= 25 and return_value > max(return_median * 2, 1):
                findings.append(
                    Finding(
                        code="MONTH_END_RETURN_SPIKE",
                        severity="high",
                        sheet_name=sheet_name,
                        row_number=index,
                        column_name=str(return_col),
                        title="Возврат перед закрытием месяца",
                        description=f"Крупный возврат {return_value:.2f} найден в конце месяца.",
                        suggested_fix="Проверьте чек возврата и причину операции перед закрытием отчетности.",
                        evidence={"return": float(return_value), "day": int(date_value.day)},
                    )
                )
    return findings[:40]


def _find_workbook_business_issues(profile: dict[str, Any]) -> list[Finding]:
    totals = profile.get("metric_totals", {})
    sales = float(totals.get("sales") or 0)
    salary = float(totals.get("salary") or 0)
    bonus = float(totals.get("bonus") or 0)
    expense = float(totals.get("expense") or 0)
    findings: list[Finding] = []
    if sales > 0 and (salary + bonus) / sales > 0.35:
        findings.append(
            Finding(
                code="PAYROLL_SHARE_HIGH",
                severity="high",
                sheet_name="Workbook",
                title="Расходы на персонал выше нормы",
                description="Фонд оплаты труда и премии превышают 35% от продаж.",
                suggested_fix="Проверьте начисления сотрудникам и сравните показатель с плановой нормой магазина.",
                evidence={"sales": sales, "salary": salary, "bonus": bonus, "ratio": round((salary + bonus) / sales, 3)},
            )
        )
    if sales > 0 and expense / sales > 0.25:
        findings.append(
            Finding(
                code="EXPENSE_SHARE_HIGH",
                severity="medium",
                sheet_name="Workbook",
                title="Расходы выше нормы",
                description="Расходы превышают 25% от продаж.",
                suggested_fix="Разбейте расходы по категориям и проверьте резкий рост аренды, персонала или закупок.",
                evidence={"sales": sales, "expense": expense, "ratio": round(expense / sales, 3)},
            )
        )
    return findings


def _find_column(columns: dict[str, Any], patterns: tuple[str, ...]) -> Any | None:
    for normalized, original in columns.items():
        if any(pattern in normalized for pattern in patterns):
            return original
    return None


def _risk_score(findings: list[Finding]) -> int:
    weights = {"info": 1, "low": 3, "medium": 8, "high": 18, "critical": 30}
    return min(100, sum(weights.get(item.severity, 1) for item in findings))


def _risk_level(score: int) -> str:
    if score >= 75:
        return "critical"
    if score >= 45:
        return "high"
    if score >= 20:
        return "medium"
    return "low"


def _summary(findings: list[Finding], risk_level: str, quality_score: int) -> str:
    if not findings:
        return f"Критичных ошибок не найдено. Оценка качества отчета: {quality_score}/100."
    critical = sum(1 for item in findings if item.severity == "critical")
    high = sum(1 for item in findings if item.severity == "high")
    fraud = sum(1 for item in findings if item.code in {"SUSPICIOUS_DISCOUNT", "MONTH_END_RETURN_SPIKE", "BONUS_LIMIT_EXCEEDED"})
    return (
        f"Найдено {len(findings)} замечаний. Критичных: {critical}, высокого риска: {high}, "
        f"fraud-сигналов: {fraud}. Уровень риска: {risk_level}. Оценка качества отчета: {quality_score}/100."
    )


def _recommendations(findings: list[Finding]) -> list[str]:
    if not findings:
        return ["Сохраните PDF проверки и приложите его к месячному закрытию."]
    codes = {item.code for item in findings}
    recommendations = []
    if "TOTAL_MISMATCH" in codes:
        recommendations.append("Сначала исправьте итоговые строки: они влияют на зарплаты, премии и финансовый результат.")
    if "BONUS_LIMIT_EXCEEDED" in codes:
        recommendations.append("Проверьте премии выше лимита и запросите основание начисления.")
    if "PAYROLL_SHARE_HIGH" in codes:
        recommendations.append("Сверьте фонд оплаты труда с плановой нормой и прошлым месяцем.")
    if "MONTH_END_RETURN_SPIKE" in codes or "SUSPICIOUS_DISCOUNT" in codes:
        recommendations.append("Проверьте подозрительные скидки и возвраты перед закрытием месяца.")
    if "REQUIRED_CELL_EMPTY" in codes:
        recommendations.append("Заполните обязательные поля, чтобы бухгалтерия могла сверить строки без ручных уточнений.")
    if "NUMERIC_OUTLIER" in codes:
        recommendations.append("Сверьте аномальные суммы с POS-системой и первичными документами.")
    return recommendations[:6]
